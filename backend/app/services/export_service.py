"""
数据导出服务
支持导出为Excel、CSV、JSON格式
"""
import csv
import json
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.models.database import Product

logger = logging.getLogger(__name__)

EXPORT_DIR = Path(__file__).resolve().parent.parent.parent.parent / "data" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# 导出字段映射
FIELD_MAP = [
    ("sku", "SKU"),
    ("title", "商品标题"),
    ("product_url", "商品链接"),
    ("image_url", "商品图片"),
    ("price", "价格（₽）"),
    ("original_price", "原价（₽）"),
    ("discount_percent", "折扣（%）"),
    ("category", "类目"),
    ("brand", "品牌"),
    ("rating", "评分"),
    ("review_count", "评论数"),
    ("monthly_sales", "月销量"),
    ("weekly_sales", "周销量"),
    ("gmv_rub", "月销售额（₽）"),
    ("paid_promo_days", "付费推广（28天参与）"),
    ("ad_cost_ratio", "广告费用占比（%）"),
    ("seller_type", "卖家类型"),
    ("seller_name", "卖家名称"),
    ("creation_date", "商品创建时间"),
    ("followers_count", "被跟数量"),
    ("follower_min_price", "被跟最低价（₽）"),
    ("follower_min_url", "被跟最低价链接"),
    ("length_cm", "长度（cm）"),
    ("width_cm", "宽度（cm）"),
    ("height_cm", "高度（cm）"),
    ("weight_g", "重量（g）"),
    ("delivery_info", "配送信息"),
    ("pdd_purchase_price", "拼多多采购价（¥）"),
    ("profit_rub", "利润（₽）"),
    ("profit_cny", "利润（¥）"),
    ("keyword", "采集关键词"),
    ("last_scraped_at", "采集时间"),
]


class ExportService:
    """数据导出服务"""

    def _product_to_dict(self, product: Product) -> dict:
        """将Product对象转为字典"""
        data = {}
        for field, _ in FIELD_MAP:
            value = getattr(product, field, None)
            if isinstance(value, datetime):
                value = value.strftime("%Y-%m-%d %H:%M:%S")
            data[field] = value
        return data

    def export_products(self, products: List[Product], format: str = "xlsx") -> str:
        """
        导出商品数据
        
        Args:
            products: 商品列表
            format: 导出格式 (xlsx/csv/json)
            
        Returns:
            导出文件路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        if format == "xlsx":
            return self._export_xlsx(products, timestamp)
        elif format == "csv":
            return self._export_csv(products, timestamp)
        elif format == "json":
            return self._export_json(products, timestamp)
        else:
            raise ValueError(f"不支持的导出格式: {format}")

    def _export_xlsx(self, products: List[Product], timestamp: str) -> str:
        """导出为Excel"""
        filepath = str(EXPORT_DIR / f"ozon_products_{timestamp}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "OZON商品数据"

        # 表头样式
        header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        headers = [label for _, label in FIELD_MAP]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        for row_idx, product in enumerate(products, 2):
            data = self._product_to_dict(product)
            for col_idx, (field, _) in enumerate(FIELD_MAP, 1):
                value = data.get(field, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 设置列宽
        column_widths = {
            "A": 15, "B": 40, "C": 50, "D": 50, "E": 12, "F": 12,
            "G": 10, "H": 25, "I": 15, "J": 8, "K": 10, "L": 10,
            "M": 10, "N": 15, "O": 18, "P": 18, "Q": 12, "R": 15,
            "S": 18, "T": 10, "U": 15, "V": 50, "W": 10, "X": 10,
            "Y": 10, "Z": 10, "AA": 15, "AB": 18, "AC": 12, "AD": 12,
            "AE": 15, "AF": 18,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # 冻结首行
        ws.freeze_panes = "A2"

        wb.save(filepath)
        logger.info(f"Excel导出完成: {filepath}, 共 {len(products)} 条数据")
        return filepath

    def _export_csv(self, products: List[Product], timestamp: str) -> str:
        """导出为CSV"""
        filepath = str(EXPORT_DIR / f"ozon_products_{timestamp}.csv")

        with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            # 写入表头
            writer.writerow([label for _, label in FIELD_MAP])
            # 写入数据
            for product in products:
                data = self._product_to_dict(product)
                writer.writerow([data.get(field, "") for field, _ in FIELD_MAP])

        logger.info(f"CSV导出完成: {filepath}, 共 {len(products)} 条数据")
        return filepath

    def _export_json(self, products: List[Product], timestamp: str) -> str:
        """导出为JSON"""
        filepath = str(EXPORT_DIR / f"ozon_products_{timestamp}.json")

        data = [self._product_to_dict(p) for p in products]
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

        logger.info(f"JSON导出完成: {filepath}, 共 {len(products)} 条数据")
        return filepath
